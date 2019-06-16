# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
<<<<<<< HEAD
from main.models import DataManager, Everything, Company
from openpyxl import load_workbook
import numpy as np
import skfuzzy as fuzz
from skfuzzy import control as ctrl
from functools import reduce
def about(request):

    # adding data from file
    '''
    wb = load_workbook('data.xlsx')
    ws = wb.active
    for row in (ws.iter_rows(min_row=2, max_col=10, max_row=308)):
        territory_name, planned_length, money, job_type_name, executor_name, year, accident_type, accident_weight, shift_weight, shift_days = row
        DataManager.add_data(territory_name.value, planned_length.value, money.value, job_type_name.value, executor_name.value, year.value, accident_type.value, accident_weight.value, shift_weight.value, shift_days.value)
    ws = wb['Организация']
    for row in (ws.iter_rows(min_row=2, max_col=2, max_row=8)):
        company_name, foundation_date = row
        DataManager.add_company(company_name.value, foundation_date.value)
    '''
    if request.user.is_authenticated:
        return render(request, 'about.html', context={'name': request.user.username})
    else:
        return render(request, 'about.html')
def count_rating(entity):
    return 100

def rating(request):
    rows = []
    for item in Everything.objects.all():
        a = (item.executor_name, 100)
        rows += [a]
    if request.user.is_authenticated:
        return render(request, 'rating.html', context={ 'name': request.user.username, 'rows':rows })
    else:
        return render(request, 'rating.html', context={ 'name': 'Войти', 'rows':rows })
=======
from main.models import DataManager
from openpyxl import Workbook, load_workbook
from main.models import Everything
from functools import reduce

def about(request):
    # adding data from file
    # wb = load_workbook('data.xlsx')
    # ws = wb.active
    # for row in (ws.iter_rows(min_row=2, max_col=10, max_row=308)):
    #     territory_name, planned_length, money, job_type_name, executor_name, year, accident_type, accident_weight, shift_weight, shift_days = row
    #     DataManager.add_data(territory_name.value, planned_length.value, money.value, job_type_name.value, executor_name.value, year.value, accident_type.value, accident_weight.value, shift_weight.value, shift_days.value)
    return render(request, template_name = 'about.html', context={ 'name': request.user.username })

def rating(request):
    def getData(orgData):
        orgProjectNames = set((list(map(lambda x: x.territory_name, orgData))))
        maxCost = max(list(map(lambda x: x.money, orgData)))
        maxAccCountInProc = max(list(map(lambda x: x.accident_weight, orgData)))

        # print('orgData = '+ str(len(orgData)))
        # print('orgProjectNames = '+ str(orgProjectNames))
        # print('maxAccInProc = '+ str(maxAccCountInProc))  

        result = []

        for orgProjectName in orgProjectNames:
            #print(orgProjectName)

            projectRecords = list(filter(lambda x: x.territory_name == orgProjectName, orgData))            
            
            inProcAccRecords = list(filter(lambda x: x.accident_type == 'В процессе', projectRecords))          
            #print('inProcAccRecords = ' + str(len(inProcAccRecords)))
            
            inResultAccRecords = list(filter(lambda x: x.accident_type == 'По результату', projectRecords))
            #print('inResultAccRecords = ' + str(len(inResultAccRecords)))

            # print(len(inResultAccRecords))
            # print(list(map(lambda x: x.accident_weight, inResultAccRecords)))

            # print(len(inProcAccRecords))
            # print(list(map(lambda x: x.accident_weight, inProcAccRecords)))

            projectData = [
                projectRecords[0].money / maxCost * 100,
                len(inProcAccRecords),
                reduce(lambda a, x: a + x, list(map(lambda x: x.accident_weight, inProcAccRecords)), 0) / 1 if len(inProcAccRecords) == 0 else len(inProcAccRecords),
                len(inResultAccRecords),
                reduce(lambda a, x: a + x, list(map(lambda x: x.accident_weight, inResultAccRecords)), 0) / 1 if len(inResultAccRecords) == 0 else len(inResultAccRecords)
            ]

            #print(projectData)
            result.append(projectData)
        
        maxInProcAccCount = max(list(map(lambda x: x[1], result)))
        maxInResultAccCount = max(list(map(lambda x: x[3], result)))
        for resultItem in result:
            resultItem[1] = resultItem[1] / (maxInProcAccCount / 100)
            resultItem[3] = resultItem[3] / (maxInResultAccCount / 100)

        return result

    from main.models import Everything 
    def normalize():
        result = []

        data = Everything.objects.all()
        tender = list(filter(lambda x: x.territory_name == 'Благоустройство 21', data))[0]
        print(tender.money)

        orgNames = list(filter(lambda x: x != '0', set(list(map(lambda x: x.executor_name, data)))))
        print(orgNames)

        for orgName in orgNames:
            print(orgName)
            orgData = list(filter(lambda x: x.executor_name == orgName, data))

            ratingData = getData(orgData)
            orgQuality = reduce(lambda a, x: a + x[0] / (x[1] * x[2] + x[3] * x[4]), ratingData, 0) / len(ratingData)

            orgExperience = len(
                list(
                    filter(lambda z: z > tender.money * 0.7,
                    set(list(map(lambda x: x.money, orgData))))
                )
            )

            result.append([orgName, orgQuality, orgExperience])
        
        # Массив массивов со структурой [Название компании, Качество, Опыт]
        print(result)
        return result

    normalize()
    return render(request, template_name='rating.html', context={'name': request.user.username})
>>>>>>> 7f31af562b2b4cdf73dbda0c652138d622b5e7b9

def form(request):
    if request.user.is_authenticated:
        return render(request, 'form.html', context={'name': request.user.username})
    else:
        return render(request, 'form.html')

def count_probability(entity):
        stag = ctrl.Antecedent(np.arange(0, 35, 1), 'stag')  # beretsya iz tablits
        k_qua = ctrl.Antecedent(np.arange(0, 100, 1), 'k_qua')  # beretsya iz tablits
        #zagr_company = ctrl.Antecedent(np.arange(0, 13, 1), 'zagr_company')  ##skolko proektov vo vremya srokov tendera
        speed = ctrl.Antecedent(np.arange(0, 115, 1), 'speed')  # beretsya iz tablits
        stoimost = ctrl.Antecedent(np.arange(1000, 50000000, 5000), 'stoimost')  ##beretsya s formi tendera
        mashtab = ctrl.Antecedent(np.arange(0, 11, 1), 'mashtab')  ##beretsya s formi tendera

        spravitsya = ctrl.Consequent(np.arange(0, 101, 1), 'spravitsya')

        stag.automf(3)
        k_qua.automf(3)
        #zagr_company.automf(3)
        speed.automf(3)
        mashtab.automf(3)
        stoimost.automf(3)

        spravitsya['low'] = fuzz.trimf(spravitsya.universe, [0, 0, 34])
        spravitsya['medium'] = fuzz.trimf(spravitsya.universe, [0, 34, 67])
        spravitsya['high'] = fuzz.trimf(spravitsya.universe, [34, 67, 101])

        rule1 = ctrl.Rule(stoimost['good'] & k_qua['good'], spravitsya['medium'])
        rule2 = ctrl.Rule(k_qua['average'] & stag['average'], spravitsya['medium'])
        rule3 = ctrl.Rule((mashtab['poor'] & stag['good']) | (k_qua['good'] & stag['good']), spravitsya['high'])
        #rule4 = ctrl.Rule(k_qua['good'] & zagr_company['average'], spravitsya['high'])
        rule5 = ctrl.Rule(stag['poor'] & k_qua['good'], spravitsya['medium'])
        rule6 = ctrl.Rule((k_qua['poor'] & speed['good']) | (stag['average'] & mashtab['good']), spravitsya['low'])
        #rule7 = ctrl.Rule((speed['average'] & stag['good']) | (stag['good'] & zagr_company['poor']),
                        #  spravitsya['medium'])
        rule8 = ctrl.Rule((speed['average'] | stag['average']), spravitsya['medium'])
        rule9 = ctrl.Rule(stoimost['average'] & stag['poor'], spravitsya['low'])
        rule10 = ctrl.Rule((mashtab['good'] | stoimost['good']) & k_qua['poor'], spravitsya['low'])
        rule11 = ctrl.Rule(mashtab['poor'] & stoimost['average'] & k_qua['good'], spravitsya['high'])

        spravitsya_ctrl = ctrl.ControlSystem([rule1, rule2, rule3, rule5, rule6, rule8, rule9, rule10, rule11])

        spravitsya_l = ctrl.ControlSystemSimulation(spravitsya_ctrl)

        import datetime
        ####
        scope = 10
        ####
        speed = entity.planned_length + entity.shift_days * entity.shift_weight
        vozrast = datetime.datetime.now().year - Company.objects.get(executor_name = entity.executor_name).foundation_date
        spravitsya_l.input['stag'] = vozrast # стаж
        spravitsya_l.input['k_qua'] = count_rating(entity) # коэфф. качества из функции (рейтинга)
        all_projects = Everything.objects.filter(executor_name=entity.executor_name).count()
        #spravitsya_l.input['zagr_company'] = 2
        # загруженность компании - сколько проектов ведет компания в промежуток времени (тендер)
        spravitsya_l.input['speed'] = speed # из табл
        spravitsya_l.input['stoimost'] = entity.money # money
        spravitsya_l.input['mashtab'] = scope # последний select из tender

        spravitsya_l.compute() #  в probability tenderа

        return spravitsya_l.output['spravitsya']

def tender(request):
    rows = []
    for item in Everything.objects.all():
        executor_contracts = Everything.objects.filter(executor_name = item.executor_name)
        sum = 0
        for i in executor_contracts:
            sum += i.money
        average_cost = sum / executor_contracts.count()
        a = (item.executor_name, count_rating(item), average_cost, float("{0:.2f}".format(count_probability(item))))
        rows += [a]
    if request.user.is_authenticated:
        return render(request, 'tender.html', context={'name': request.user.username, 'rows': rows})
    else:
        return render(request, 'tender.html', context={ 'name': 'Войти', 'rows':rows })

def login_page(request):
    if request.method=='POST' and request.POST['action']=='login':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                return redirect('/')
            else:
                return HttpResponse('К сожалению, аккаунт не существует. <button><a href="/">Вернуться назад</a></button>')
        else:
            return redirect('/login_page', permanent=False)
    elif request.method=='POST' and request.POST['action']=='logout':
        logout(request)
        return render(request,'login_page.html',context={})
    else:#if request.method=='GET':
        if request.user.is_authenticated:
            return render(request,'login_page.html', context={ 'name': request.user.username })
        else:
            return render(request,'login_page.html')

